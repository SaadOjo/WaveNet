from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(SCRIPT_DIR) in sys.path:
    sys.path.remove(str(SCRIPT_DIR))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.io import project_root

METRICS = ["Accuracy", "Precision", "Recall", "F1-score", "ROC-AUC score"]
DISPLAY_CLASSICAL_MODELS = ["Random Forest", "XGBoost"]
HIDDEN_MODELS = {"Support Vector Machine"}
MODEL_ORDER = [
    "Random Forest",
    "XGBoost",
    "UniTS",
    "WaveNet",
    "SmallWaveNetTL",
    "WaveNetTL",
    "WaveNet_TL_CIFAR",
    "WaveNet_TL_Resnet_Base",
    "ResNeXt50_TransferLearning",
    "ResNeXt50_TransferLearning-v2",
    "ResNeXt101_TransferLearning",
    "ViT_B_16_TransferLearning",
]

MONO_FONT = "JetBrains Mono"
FALLBACK_FONT_SIZE = 11


def latest_run_dirs(runs_root: Path) -> dict[str, Path]:
    latest = {}
    if not runs_root.exists():
        return latest

    for model_dir in sorted(p for p in runs_root.iterdir() if p.is_dir()):
        run_dirs = sorted(p for p in model_dir.iterdir() if p.is_dir())
        if run_dirs:
            latest[model_dir.name] = run_dirs[-1]
    return latest


def load_dataset_metadata(metadata_path: Path) -> dict[str, dict]:
    if not metadata_path.exists():
        return {}

    rows = json.loads(metadata_path.read_text())
    metadata = {}
    for row in rows:
        export_ts = row.get("export_ts", {})
        output_dir = export_ts.get("output_dir", "")
        dataset_folder = Path(output_dir).name
        if not dataset_folder:
            continue

        ratio_values = export_ts.get("churner_to_nonchurner_rate") or row.get("churner_to_nonchurner_rate")
        if isinstance(ratio_values, list) and len(ratio_values) == 2:
            ratio = f"{ratio_values[0]}:{ratio_values[1]}"
        else:
            ratio = "?"

        max_customers = export_ts.get("max_customers")
        if max_customers is None:
            size = "full"
        else:
            size = f"max{max_customers}"

        metadata[dataset_folder] = {
            "data_no": str(row.get("data_no", "?")),
            "source": row.get("data_name", "?"),
            "ratio": ratio,
            "size": size,
        }
    return metadata


def pretty_dataset_label(dataset_name: str, metadata_map: dict[str, dict]) -> str:
    meta = metadata_map.get(dataset_name)
    if meta:
        return f"data{meta['data_no']} ({meta['source']}, {meta['ratio']}, {meta['size']})"
    return dataset_name


def load_latest_results(runs_root: Path) -> list[dict]:
    records = []
    for script_name, run_dir in latest_run_dirs(runs_root).items():
        results_path = run_dir / "results.json"
        config_path = run_dir / "config.json"
        if not results_path.exists():
            continue

        rows = json.loads(results_path.read_text())
        config = json.loads(config_path.read_text()) if config_path.exists() else {}
        run_id = run_dir.name

        for row in rows:
            enriched = dict(row)
            enriched["script_name"] = script_name
            enriched["run_id"] = run_id
            enriched["run_dir"] = str(run_dir)
            enriched["target_ids"] = config.get("target_ids", [])
            records.append(enriched)
    return records


def sort_models(models: list[str]) -> list[str]:
    order_index = {name: idx for idx, name in enumerate(MODEL_ORDER)}
    return sorted(models, key=lambda name: (order_index.get(name, 10_000), name))


def build_dataset_view(records: list[dict]):
    dataset_metric_model = defaultdict(lambda: defaultdict(dict))
    datasets = set()
    models = set()

    for row in records:
        dataset = row["data_name"]
        model = row["model"]
        if model in HIDDEN_MODELS:
            continue
        datasets.add(dataset)
        models.add(model)
        for metric in METRICS:
            value = row.get(metric)
            if value is not None:
                dataset_metric_model[dataset][metric][model] = value

    datasets = sorted(datasets)
    models = sort_models(list(models))
    return datasets, models, dataset_metric_model


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(width + 2, 28)


def is_numeric(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def build_workbook(records: list[dict], output_path: Path, metadata_map: dict[str, dict]):
    if not records:
        raise SystemExit("No run results found under runs/ to export.")

    datasets, display_models, dataset_metric_model = build_dataset_view(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "by_dataset"

    ordered_display_models = [
        *[model for model in DISPLAY_CLASSICAL_MODELS if model in display_models],
        *[model for model in display_models if model not in DISPLAY_CLASSICAL_MODELS],
    ]
    header = ["Dataset", "Metric", *ordered_display_models]
    ws.append(header)

    header_fill = PatternFill("solid", fgColor="4A4A4A")
    dataset_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(name=MONO_FONT, size=FALLBACK_FONT_SIZE, color="FFFFFF", bold=True)
    dataset_font = Font(name=MONO_FONT, size=FALLBACK_FONT_SIZE, color="222222", bold=True)
    metric_font = Font(name=MONO_FONT, size=FALLBACK_FONT_SIZE, color="222222", bold=True)
    body_font = Font(name=MONO_FONT, size=FALLBACK_FONT_SIZE, color="222222")
    best_font = Font(name=MONO_FONT, size=FALLBACK_FONT_SIZE, color="222222", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    medium = Side(style="medium", color="7F7F7F")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 2
    for dataset in datasets:
        pretty_label = pretty_dataset_label(dataset, metadata_map)
        start_row = current_row
        for metric in METRICS:
            metric_values = dataset_metric_model[dataset].get(metric, {})
            row = [pretty_label, metric]
            for model in ordered_display_models:
                row.append(metric_values.get(model))
            ws.append(row)
            current_row += 1

        end_row = current_row - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        dataset_cell = ws.cell(row=start_row, column=1)
        dataset_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dataset_cell.fill = dataset_fill
        dataset_cell.font = dataset_font

        for row_idx in range(start_row, end_row + 1):
            ws.cell(row=row_idx, column=2).font = metric_font
            ws.cell(row=row_idx, column=2).fill = white_fill
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")

            numeric_values = []
            row_cells = []
            for col_idx in range(3, 3 + len(ordered_display_models)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.fill = white_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_numeric(cell.value):
                    cell.number_format = "0.0000"
                    numeric_values.append(float(cell.value))
                row_cells.append(cell)

            if numeric_values:
                best_value = max(numeric_values)
                for cell in row_cells:
                    if is_numeric(cell.value) and abs(float(cell.value) - best_value) <= 1e-12:
                        cell.font = best_font

        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, 3 + len(ordered_display_models)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, 3 + len(ordered_display_models)):
            ws.cell(row=end_row, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=medium)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    autosize(ws)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14

    filtered_records = [row for row in records if row.get("model") not in HIDDEN_MODELS]

    raw = wb.create_sheet("latest_raw")
    raw_headers = ["run_id", "script_name", "data_id", "data_name", "model", *METRICS, "run_dir"]
    raw.append(raw_headers)
    for cell in raw[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sorted(filtered_records, key=lambda r: (r["data_name"], r["model"])):
        raw.append([row.get(key) for key in raw_headers])

    for row in raw.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.fill = white_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if isinstance(cell.value, (float, int)) and 6 <= cell.column <= 10:
                cell.number_format = "0.0000"
    raw.freeze_panes = "A2"
    autosize(raw)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Export the latest run from each model into a formatted Excel report.")
    parser.add_argument("--output", default=None, help="Output .xlsx path. Defaults to results/latest_model_comparison.xlsx")
    args = parser.parse_args()

    root = project_root()
    output_path = Path(args.output) if args.output else root / "results" / "latest_model_comparison.xlsx"
    records = load_latest_results(root / "runs")
    metadata_map = load_dataset_metadata(root / "data" / "dataset_config.json")
    build_workbook(records, output_path, metadata_map)
    print(f"Wrote Excel report to {output_path}")


if __name__ == "__main__":
    main()
