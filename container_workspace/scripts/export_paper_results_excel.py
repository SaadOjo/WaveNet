from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(SCRIPT_DIR) in sys.path:
    sys.path.remove(str(SCRIPT_DIR))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.io import project_root

METRICS = ["Accuracy", "Precision", "Recall", "F1-score", "ROC-AUC score"]
MONO_FONT = "JetBrains Mono"
FONT_SIZE = 11

MODEL_RUN_SPECS = [
    {
        "run_stem": "random_forest",
        "model_name": "Random Forest",
        "display_name": "Random Forest",
    },
    {
        "run_stem": "xgboost",
        "model_name": "XGBoost",
        "display_name": "XGBoost",
    },
    {
        "run_stem": "small_wavenet_v3_tl",
        "model_name": "SmallWaveNetV3_TL",
        "display_name": "SmallWaveNetTL",
    },
    {
        "run_stem": "wavelet_vit",
        "model_name": "ViT_B_16_TransferLearning",
        "display_name": "ViT",
    },
]

TARGET_DATA_IDS = ["10", "11", "12", "13"]


def latest_run_dir(runs_root: Path, run_stem: str) -> Path:
    model_root = runs_root / run_stem
    run_dirs = sorted(p for p in model_root.iterdir() if p.is_dir())
    if not run_dirs:
        raise FileNotFoundError(f"No run dirs found for {run_stem}")
    return run_dirs[-1]


def load_dataset_metadata(metadata_path: Path) -> dict[str, dict]:
    if not metadata_path.exists():
        return {}

    rows = json.loads(metadata_path.read_text())
    metadata = {}
    for row in rows:
        export_ts = row.get("export_ts", {})
        output_dir = export_ts.get("output_dir", "")
        dataset_folder = Path(output_dir).name
        if not dataset_folder:
            continue

        ratio_values = export_ts.get("churner_to_nonchurner_rate") or row.get("churner_to_nonchurner_rate")
        if isinstance(ratio_values, list) and len(ratio_values) == 2:
            ratio = f"{ratio_values[0]}:{ratio_values[1]}"
        else:
            ratio = "?"

        max_customers = export_ts.get("max_customers")
        size = "full" if max_customers is None else f"max{max_customers}"
        metadata[dataset_folder] = {
            "data_no": str(row.get("data_no", "?")),
            "source": row.get("data_name", "?"),
            "ratio": ratio,
            "size": size,
        }
    return metadata


def pretty_dataset_label(dataset_name: str, metadata_map: dict[str, dict]) -> str:
    meta = metadata_map.get(dataset_name)
    if meta:
        return f"data{meta['data_no']} ({meta['source']}, {meta['ratio']}, {meta['size']})"
    return dataset_name


def load_selected_results(runs_root: Path):
    selected = {}
    raw_rows = []
    for spec in MODEL_RUN_SPECS:
        run_dir = latest_run_dir(runs_root, spec["run_stem"])
        rows = json.loads((run_dir / "results.json").read_text())
        matched_rows = [
            row
            for row in rows
            if row.get("model") == spec["model_name"] and row.get("data_id") in TARGET_DATA_IDS
        ]
        by_data_id = {row["data_id"]: row for row in matched_rows}
        selected[spec["display_name"]] = {
            "run_stem": spec["run_stem"],
            "model_name": spec["model_name"],
            "run_dir": run_dir,
            "rows": by_data_id,
        }
        for row in matched_rows:
            raw = dict(row)
            raw["display_model"] = spec["display_name"]
            raw["run_stem"] = spec["run_stem"]
            raw["run_dir"] = str(run_dir)
            raw_rows.append(raw)
    return selected, raw_rows


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(width + 2, 30)


def build_workbook(output_path: Path, selected: dict, raw_rows: list[dict], metadata_map: dict[str, dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "focused_ic1"

    model_headers = [spec["display_name"] for spec in MODEL_RUN_SPECS]
    ws.append(["Dataset", "Metric", *model_headers])

    header_fill = PatternFill("solid", fgColor="4A4A4A")
    dataset_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(name=MONO_FONT, size=FONT_SIZE, color="FFFFFF", bold=True)
    dataset_font = Font(name=MONO_FONT, size=FONT_SIZE, color="222222", bold=True)
    metric_font = Font(name=MONO_FONT, size=FONT_SIZE, color="222222", bold=True)
    body_font = Font(name=MONO_FONT, size=FONT_SIZE, color="222222")
    best_font = Font(name=MONO_FONT, size=FONT_SIZE, color="222222", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    medium = Side(style="medium", color="7F7F7F")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 2
    for data_id in TARGET_DATA_IDS:
        sample_row = None
        for header in model_headers:
            sample_row = selected[header]["rows"].get(data_id)
            if sample_row is not None:
                break
        if sample_row is None:
            continue

        dataset_name = sample_row["data_name"]
        pretty_label = pretty_dataset_label(dataset_name, metadata_map)
        start_row = current_row
        for metric in METRICS:
            row = [pretty_label, metric]
            values = []
            for header in model_headers:
                value = selected[header]["rows"].get(data_id, {}).get(metric)
                row.append(value)
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    values.append(float(value))
            ws.append(row)
            row_idx = current_row
            for col_idx in range(3, 3 + len(model_headers)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.fill = white_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "0.0000"
            if values:
                best_value = max(values)
                for col_idx in range(3, 3 + len(model_headers)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        if abs(float(cell.value) - best_value) <= 1e-12:
                            cell.font = best_font
            current_row += 1

        end_row = current_row - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        dataset_cell = ws.cell(row=start_row, column=1)
        dataset_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dataset_cell.fill = dataset_fill
        dataset_cell.font = dataset_font

        for row_idx in range(start_row, end_row + 1):
            metric_cell = ws.cell(row=row_idx, column=2)
            metric_cell.font = metric_font
            metric_cell.fill = white_fill
            metric_cell.alignment = Alignment(horizontal="left", vertical="center")
            for col_idx in range(1, 3 + len(model_headers)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx in range(1, 3 + len(model_headers)):
            ws.cell(row=end_row, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=medium)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    autosize(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    raw = wb.create_sheet("raw_selected")
    raw_headers = ["display_model", "run_stem", "data_id", "data_name", "model", *METRICS, "run_dir"]
    raw.append(raw_headers)
    for cell in raw[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sorted(raw_rows, key=lambda r: (r["data_id"], r["display_model"])):
        raw.append([row.get(key) for key in raw_headers])
    for row in raw.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.fill = white_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if isinstance(cell.value, (int, float)) and 6 <= cell.column <= 10:
                cell.number_format = "0.0000"
    raw.freeze_panes = "A2"
    autosize(raw)

    notes = wb.create_sheet("notes")
    notes_rows = [
        ["Focused comparison scope"],
        ["This workbook includes only the common-coverage IC1 datasets: data10, data11, data12, data13."],
        ["Model mapping"],
        ["The SmallWaveNetTL column is populated from the reproducible refactored run SmallWaveNetV3_TL."],
        ["ViT column refers to the refactored ViT_B_16_TransferLearning baseline (ImageNet-pretrained, no churn pretraining)."],
    ]
    for row in notes_rows:
        notes.append(row)
    notes.column_dimensions["A"].width = 120
    for idx, row in enumerate(notes.iter_rows(min_row=1, max_row=notes.max_row), start=1):
        for cell in row:
            cell.font = header_font if idx in {1, 3} else body_font
            cell.fill = header_fill if idx in {1, 3} else white_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    root = project_root()
    repo_root = root.parent
    selected, raw_rows = load_selected_results(root / "runs")
    metadata_map = load_dataset_metadata(root / "data" / "dataset_config.json")
    output_path = repo_root / "paper_results" / "focused_ic1_results.xlsx"
    build_workbook(output_path, selected, raw_rows, metadata_map)
    print(f"Wrote focused workbook to {output_path}")


if __name__ == "__main__":
    main()
